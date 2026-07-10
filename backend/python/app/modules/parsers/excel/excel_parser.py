import asyncio
import io
import json
import logging
import os
import re
from datetime import datetime
from typing import Any

from langchain_core.language_models.chat_models import BaseChatModel
from langchain_core.messages import AIMessage, HumanMessage
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from tenacity import (
    retry,
    stop_after_attempt,
    wait_exponential,
)

from app.models.blocks import (
    Block,
    BlockGroup,
    BlockGroupChildren,
    BlocksContainer,
    BlockType,
    DataFormat,
    GroupType,
    TableMetadata,
)
from app.modules.parsers.excel.prompt_template import (
    ExcelHeaderDetection,
    TableHeaders,
    excel_header_detection_prompt,
    excel_header_generation_prompt,
    row_text_prompt,
    row_text_prompt_for_csv,
    sheet_summary_prompt,
    table_summary_prompt,
)
from app.utils.aimodels import coerce_message_content_to_text
from app.utils.indexing_helpers import format_rows_with_index, generate_simple_row_text
from app.utils.streaming import (
    invoke_with_row_descriptions_and_reflection,
    invoke_with_structured_output_and_reflection,
)

# Module-level constants for Excel processing (mirror CSV parser)
NUM_SAMPLE_ROWS = 5  # Number of representative sample rows to select for header generation
MAX_HEADER_GENERATION_ROWS = 10  # Maximum number of rows to use for header generation
MAX_HEADER_DETECTION_ROWS = 4  # Check first 4 rows for multi-row headers
MIN_ROWS_FOR_HEADER_ANALYSIS = 1  # Minimum number of rows required for header analysis
MAX_HEADER_COUNT_RETRIES = 2  # Maximum retries when LLM returns wrong header count
EXCEL_HEADER_GENERATION_SAMPLE_SCAN_LIMIT = max(50, int(os.getenv("EXCEL_HEADER_GENERATION_SAMPLE_SCAN_LIMIT", "500")))
EXCEL_MAX_TABLE_ROWS_TO_INDEX = max(1, int(os.getenv("EXCEL_MAX_TABLE_ROWS_TO_INDEX", "20000")))


# Built-in Excel date format codes mapping
BUILTIN_DATE_FORMATS = {
    14: "mm/dd/yyyy",
    15: "d-mmm-yy",
    16: "d-mmm",
    17: "mmm-yy",
    18: "h:mm AM/PM",
    19: "h:mm:ss AM/PM",
    20: "h:mm",
    21: "h:mm:ss",
    22: "m/d/yy h:mm",
}

# Common Excel datetime format whitelist - maps Excel format to (Python strftime format, needs_leading_zero_strip)
# Covers 80-90% of real-world usage with simple, maintainable mappings
# Format: "excel_format": ("python_format", "strip_pattern")
# strip_pattern: "d"=day, "m"=month, "h"=hour, "dm"=day+month, etc.
COMMON_FORMAT_WHITELIST = {
    # Date-only formats (no ambiguity - mm is always month)
    "mm/dd/yyyy": ("%m/%d/%Y", ""),
    "dd/mm/yyyy": ("%d/%m/%Y", ""),
    "yyyy-mm-dd": ("%Y-%m-%d", ""),
    "mm-dd": ("%m-%d", ""),
    "mm/yy": ("%m/%y", ""),
    "mm-yyyy": ("%m-%Y", ""),
    "dd-mmm-yy": ("%d-%b-%y", ""),
    "d-mmm-yy": ("%d-%b-%y", "d"),
    "mmm-yyyy": ("%b-%Y", ""),
    "mmm-yy": ("%b-%y", ""),
    "d-mmm": ("%d-%b", "d"),
    "mmmm yyyy": ("%B %Y", ""),
    "mmmm d, yyyy": ("%B %d, %Y", "d"),

    # Time-only formats (no ambiguity - mm is always minute with colons)
    "h:mm": ("%H:%M", "h"),
    "hh:mm": ("%H:%M", ""),
    "hh:mm:ss": ("%H:%M:%S", ""),
    "h:mm:ss": ("%H:%M:%S", "h"),
    "h:mm AM/PM": ("%I:%M %p", "h"),
    "hh:mm AM/PM": ("%I:%M %p", ""),
    "h:mm:ss AM/PM": ("%I:%M:%S %p", "h"),
    "hh:mm:ss AM/PM": ("%I:%M:%S %p", ""),

    # Combined date-time formats (first mm=month, second mm=minute)
    "mm/dd/yyyy h:mm": ("%m/%d/%Y %H:%M", "h"),
    "mm/dd/yyyy hh:mm": ("%m/%d/%Y %H:%M", ""),
    "dd/mm/yyyy h:mm": ("%d/%m/%Y %H:%M", "h"),
    "dd/mm/yyyy hh:mm": ("%d/%m/%Y %H:%M", ""),
    "dd-mm-yyyy hh:mm": ("%d-%m-%Y %H:%M", ""),
    "mm/dd/yyyy hh:mm:ss": ("%m/%d/%Y %H:%M:%S", ""),
    "mm/dd/yy, h:mm:ss": ("%m/%d/%y, %H:%M:%S", "h"),
    "m/d/yy h:mm": ("%m/%d/%y %H:%M", "dmh"),
    "m/d/yyyy h:m": ("%m/%d/%Y %H:%M", "dmh"),
    "dd-mmm-yy hh:mm": ("%d-%b-%y %H:%M", ""),
    "mmm dd, yyyy h:mm AM/PM": ("%b %d, %Y %I:%M %p", "dh"),
    "mm/dd/yyyy h:mm AM/PM": ("%m/%d/%Y %I:%M %p", "h"),

    # Edge cases that are uncommon but appear in tests
    "mm:ss": ("%M:%S", ""),  # Minutes:seconds format (no hours)
    "hh : mm : ss": ("%H : %M : %S", ""),  # Spaces around colons
}


def _resolve_builtin_format(number_format: str) -> str:
    """
    Convert built-in Excel format codes (14-22) to format strings.

    Args:
        number_format: Excel format code or string

    Returns:
        Format string (either from BUILTIN_DATE_FORMATS or original)
    """
    if number_format.isdigit():
        code = int(number_format)
        return BUILTIN_DATE_FORMATS.get(code, number_format)
    return number_format


def _strip_leading_zeros(formatted: str, strip_pattern: str) -> str:
    """
    Strip leading zeros from formatted datetime string based on pattern.

    Args:
        formatted: The formatted datetime string
        strip_pattern: String indicating what to strip ("d"=day, "m"=month, "h"=hour, or combinations)

    Returns:
        String with leading zeros stripped as specified
    """
    if not strip_pattern:
        return formatted

    # Strip leading zeros from day (e.g., "05" -> "5")
    if 'd' in strip_pattern:
        formatted = re.sub(r'(?<![0-9])0([1-9])(?=[^0-9]|$)', r'\1', formatted)

    # Strip leading zeros from month (e.g., "03" -> "3")
    if 'm' in strip_pattern:
        # Strip all matching occurrences (consistent with day/hour handling)
        # Note: In practice, when 'm' is in strip_pattern, 'd' is usually also present,
        # and the day pattern above will have already stripped matching patterns
        formatted = re.sub(r'(?<![0-9])0([1-9])', r'\1', formatted)

    # Strip leading zeros from hours (e.g., "02:30" -> "2:30")
    if 'h' in strip_pattern:
        formatted = re.sub(r'(?<![0-9])0([1-9])(?=:)', r'\1', formatted)

    return formatted


def _apply_post_processing(formatted: str, original_format: str, strip_pattern: str = "") -> str:
    """
    Apply Excel-specific formatting nuances.

    Args:
        formatted: The formatted datetime string
        original_format: The original Excel format string
        strip_pattern: Pattern for stripping leading zeros

    Returns:
        Post-processed formatted string
    """
    # Strip leading zeros if needed
    formatted = _strip_leading_zeros(formatted, strip_pattern)

    # Lowercase month abbreviations if original format had lowercase 'mmm'
    if 'mmm' in original_format.lower() and 'mmmm' not in original_format.lower():
        for month in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']:
            if month in formatted:
                formatted = formatted.replace(month, month.lower(), 1)
                break

    return formatted


def _resolve_ambiguous_format(format_str: str) -> str:
    """
    Handle formats not in whitelist using simple regex heuristics.

    Key insight:
    - 'mm' or 'm' adjacent to ':' (with colon) → minute
    - 'mm' or 'm' in date context (no colons nearby) → month
    - For combined formats, mark time-context 'mm'/'m' before replacing

    Args:
        format_str: Excel format string

    Returns:
        Python strftime format string
    """
    python_format = format_str

    # 1. Handle month/day names first (no ambiguity)
    python_format = python_format.replace("mmmm", "%B")
    python_format = python_format.replace("dddd", "%A")
    python_format = python_format.replace("mmm", "%b")
    python_format = python_format.replace("ddd", "%a")

    # 2. Handle years
    python_format = python_format.replace("yyyy", "%Y")
    python_format = python_format.replace("yy", "%y")

    # 3. Detect if format has time component
    has_time = bool(re.search(r'(?<![a-z])h+(?![a-z])', format_str, re.IGNORECASE))
    has_am_pm = 'am/pm' in format_str.lower() or 'a/p' in format_str.lower()

    # 4. Handle AM/PM
    python_format = re.sub(r'AM/PM', '%p', python_format, flags=re.IGNORECASE)
    python_format = re.sub(r'A/P', '%p', python_format, flags=re.IGNORECASE)

    # 5. Handle hours (context-sensitive: 12-hour vs 24-hour based on AM/PM)
    if 'hh' in python_format.lower():
        if has_am_pm:
            python_format = re.sub(r'hh', '%I', python_format, flags=re.IGNORECASE)
        else:
            python_format = re.sub(r'hh', '%H', python_format, flags=re.IGNORECASE)

    if re.search(r'(?<!%)h(?![a-zA-Z])', python_format, flags=re.IGNORECASE):
        if has_am_pm:
            python_format = re.sub(r'(?<!%)h(?![a-zA-Z])', '%I', python_format, flags=re.IGNORECASE)
        else:
            python_format = re.sub(r'(?<!%)h(?![a-zA-Z])', '%H', python_format, flags=re.IGNORECASE)

    # 6. Handle seconds
    python_format = python_format.replace("ss", "%S")
    python_format = re.sub(r'(?<!%)s(?![a-zA-Z])', '%S', python_format)

    # 7. Handle ambiguous 'mm' using context - IMPORTANT: Mark time-context mm FIRST
    if has_time:
        # Special case: mm:ss format (no hours) - mm is minutes
        if not bool(re.search(r'[hH%]', python_format)):
            python_format = python_format.replace('mm', '%M')
        else:
            # Mark time-context mm with placeholders to preserve them
            # Match :mm or mm: (with optional spaces)
            python_format = re.sub(r':\s*mm(?![m:])', ':__MINUTE_MM__', python_format)
            python_format = re.sub(r'(?<![m:])mm\s*:', '__MINUTE_MM__:', python_format)

            # Now replace any remaining mm (date context) with month
            python_format = python_format.replace('mm', '%m')

            # Replace placeholders with minute format
            python_format = python_format.replace('__MINUTE_MM__', '%M')
    else:
        # Date-only: mm → month
        python_format = python_format.replace('mm', '%m')

    # 8. Handle 'dd' - day with leading zero
    python_format = python_format.replace("dd", "%d")

    # 9. Handle single 'd' - day (will strip leading zero in post-processing)
    if re.search(r'(?<!%)d(?![a-zA-Z])', python_format):
        python_format = re.sub(r'(?<!%)d(?![a-zA-Z])', '%d', python_format)

    # 10. Handle single 'm' - month or minute (context-dependent, will strip leading zero in post-processing)
    if re.search(r'(?<!%)m(?![a-zA-Z])', python_format):
        if has_time:
            # Mark time-context m with placeholders
            python_format = re.sub(r':\s*m(?![ma-zA-Z])', ':__MINUTE_M__', python_format)
            python_format = re.sub(r'(?<![m:])m\s*:', '__MINUTE_M__:', python_format)

            # Remaining 'm' is month
            python_format = re.sub(r'(?<!%)m(?![a-zA-Z_])', '%m', python_format)

            # Replace placeholders
            python_format = python_format.replace('__MINUTE_M__', '%M')
        else:
            # Date-only: m → month
            python_format = re.sub(r'(?<!%)m(?![a-zA-Z])', '%m', python_format)

    return python_format


def format_excel_datetime(dt_value: datetime | str | int | float | None, number_format: str) -> str | int | float | None:
    """
    Apply Excel number format to datetime value using whitelist-based approach.

    Supported formats:
    - Built-in Excel codes: 14-22 (mm/dd/yyyy, h:mm, etc.)
    - Common date formats: mm/dd/yyyy, dd/mm/yyyy, yyyy-mm-dd, m/d/yy, etc.
    - Common time formats: h:mm, hh:mm:ss, h:mm AM/PM, etc.
    - Mixed formats: mm/dd/yyyy h:mm, m/d/yy h:mm
    - Month names: mmm-yyyy, dd-mmm-yy, mmmm yyyy

    Unsupported/edge case formats fall back to ISO format.

    Args:
        dt_value: The cell value (may or may not be a datetime)
        number_format: The Excel number format code (e.g., "mmm-yyyy", "dd-mmm-yy")

    Returns:
        Formatted string if datetime with valid format, otherwise original value
    """
    # Early returns for non-datetime values
    if not isinstance(dt_value, datetime):
        return dt_value

    if not number_format or number_format == "General":
        return dt_value.isoformat()

    try:
        # Step 1: Resolve built-in format codes (14-22) to format strings
        format_str = _resolve_builtin_format(number_format)

        # Step 2: Try whitelist first (covers 80-90% of cases with simple lookup)
        if format_str in COMMON_FORMAT_WHITELIST:
            python_fmt, strip_pattern = COMMON_FORMAT_WHITELIST[format_str]
            formatted = dt_value.strftime(python_fmt)
            return _apply_post_processing(formatted, format_str, strip_pattern)

        # Step 3: Try simple regex resolution for non-whitelisted formats
        python_fmt = _resolve_ambiguous_format(format_str)
        formatted = dt_value.strftime(python_fmt)
        return _apply_post_processing(formatted, format_str, "")

    except Exception:
        # Fallback to ISO for truly unsupported formats
        return dt_value.isoformat()


class ExcelParser:
    def __init__(self, logger: logging.Logger) -> None:
        self.logger = logger
        self.workbook = None
        self.file_binary = None

        # Store prompts
        self.sheet_summary_prompt = sheet_summary_prompt
        self.table_summary_prompt = table_summary_prompt
        self.row_text_prompt = row_text_prompt

        # Configure retry parameters
        self.max_retries = 3
        self.min_wait = 1  # seconds
        self.max_wait = 10  # seconds

    def load_workbook_from_binary(self, file_binary: bytes) -> None:
        """Load workbook from binary (no LLM calls).

        This is the first phase of Excel processing - pure parsing without LLM calls.
        """
        self.logger.info("Loading workbook from binary data")
        self.file_binary = file_binary
        if self.file_binary:
            self.workbook = load_workbook(io.BytesIO(file_binary), data_only=True)
            self.logger.info(f"Workbook loaded successfully with {len(self.workbook.sheetnames)} sheets: {self.workbook.sheetnames}")

    async def create_blocks(self, llm: BaseChatModel) -> BlocksContainer:
        """Create blocks from loaded workbook (involves LLM calls).

        This is the second phase - involves LLM calls for table summaries and row descriptions.
        Must call load_workbook_from_binary() first.
        """
        self.logger.info("Starting block creation from workbook with LLM")
        try:
            result = await self.get_blocks_from_workbook(llm)
            self.logger.info(f"Block creation completed. Generated {len(result.blocks)} blocks and {len(result.block_groups)} block groups")
            return result
        finally:
            if self.workbook:
                self.logger.info("Closing workbook")
                self.workbook.close()


    def _json_default(self, obj: object) -> str:
        if isinstance(obj, datetime):
            return obj.isoformat()
        return str(obj)

    def _process_sheet(self, sheet: Worksheet) -> dict[str, list[list[dict[str, Any]]]]:
        """Process individual sheet and extract cell data"""
        try:
            self.logger.info(f"Processing sheet: {sheet.title}")
            sheet_data = {"headers": [], "data": []}

            # Extract headers from first row
            first_row = next(sheet.iter_rows(min_row=1, max_row=1))
            sheet_data["headers"] = [cell.value for cell in first_row]
            self.logger.info(f"Extracted {len(sheet_data['headers'])} headers from first row")

            # Start from second row
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), 2):
                row_data = []

                for col_idx, cell in enumerate(row, 1):
                    # Handle merged cells
                    if isinstance(cell, MergedCell):
                        cell_data = {
                            "value": None,  # Merged cells don't contain values
                            "header": (
                                sheet_data["headers"][col_idx - 1]
                                if col_idx - 1 < len(sheet_data["headers"])
                                else None
                            ),
                            "row": row_idx,
                            "column": col_idx,
                            # Use utility function instead
                            "column_letter": get_column_letter(col_idx),
                            "coordinate": f"{get_column_letter(col_idx)}{row_idx}",
                            "data_type": "merged",
                            "style": {"font": {}, "fill": {}, "alignment": {}},
                        }
                    else:
                        cell_data = {
                            "value": cell.value,
                            "header": (
                                sheet_data["headers"][col_idx - 1]
                                if col_idx - 1 < len(sheet_data["headers"])
                                else None
                            ),
                            "row": row_idx,
                            "column": col_idx,
                            "column_letter": cell.column_letter,
                            "coordinate": cell.coordinate,
                            "data_type": cell.data_type,
                            "style": {
                                "font": {
                                    "bold": cell.font.bold,
                                    "italic": cell.font.italic,
                                    "size": cell.font.size,
                                    "color": (
                                        cell.font.color.rgb if cell.font.color else None
                                    ),
                                },
                                "fill": {
                                    "background_color": (
                                        cell.fill.start_color.rgb
                                        if cell.fill.start_color
                                        else None
                                    )
                                },
                                "alignment": {
                                    "horizontal": cell.alignment.horizontal,
                                    "vertical": cell.alignment.vertical,
                                },
                            },
                        }

                        # Add formula if present
                        if cell.data_type == "f":
                            cell_data["formula"] = cell.value

                    row_data.append(cell_data)

                sheet_data["data"].append(row_data)

            self.logger.info(f"Processed {len(sheet_data['data'])} data rows from sheet: {sheet.title}")
            return sheet_data

        except Exception as e:
            self.logger.error(f"Error processing sheet {sheet.title}: {e}", exc_info=True)
            raise

    async def find_tables(self, sheet: Worksheet, llm: BaseChatModel) -> list[dict[str, Any]]:
        """Find and process all tables in a sheet with LLM-based header detection/generation"""
        try:
            self.logger.info(f"Finding tables in sheet: {sheet.title}")
            tables = []
            visited_ranges: list[tuple[int, int, int, int]] = []

            # Pre-scan: build a set of non-empty cell positions and find the true sheet
            # extent. openpyxl frequently reports inflated max_row/max_col (e.g. 98k rows)
            # for sheets that only have ~1k rows of actual data because empty-but-styled
            # rows are counted. Iterating sheet._cells (the internal non-empty cell dict)
            # is 50-100x faster than iter_rows() for such sheets.
            # NOTE: sheet._cells is a private openpyxl implementation detail (tested with
            # openpyxl==3.1.5). It may change or be removed in future versions. The
            # getattr fallback below ensures graceful degradation to iter_rows() if the
            # attribute no longer exists.
            _raw_cells = getattr(sheet, '_cells', None)
            if _raw_cells is not None:
                non_empty_positions: set[tuple[int, int]] = {
                    (r, c) for (r, c), _cell in _raw_cells.items() if _cell.value is not None
                }
            else:
                non_empty_positions = {
                    (cell.row, cell.column)
                    for _row in sheet.iter_rows()
                    for cell in _row
                    if cell.value is not None
                }

            if not non_empty_positions:
                self.logger.info(f"Sheet '{sheet.title}' has no data, skipping")
                return []

            effective_max_row: int = max(r for r, _ in non_empty_positions)
            effective_max_col: int = max(c for _, c in non_empty_positions)

            if effective_max_row < sheet.max_row or effective_max_col < sheet.max_column:
                self.logger.info(
                    f"Sheet '{sheet.title}': effective dimensions {effective_max_row}\u00d7{effective_max_col} "
                    f"(openpyxl reported {sheet.max_row}\u00d7{sheet.max_column})"
                )

            def get_visited_range(row: int, col: int) -> tuple[int, int, int, int] | None:
                for start_row, end_row, start_col, end_col in visited_ranges:
                    if start_row <= row <= end_row and start_col <= col <= end_col:
                        return (start_row, end_row, start_col, end_col)
                return None

            async def get_table(start_row: int, start_col: int) -> dict[str, Any]:
                """Extract a table starting from (start_row, start_col) with intelligent header detection."""
                self.logger.info(f"Extracting table starting at row={start_row}, col={start_col}")
                # Step 1: Find table boundaries (max_row, max_col)
                max_col = start_col
                for col in range(start_col, effective_max_col + 1):
                    has_data = False
                    for r in range(start_row, effective_max_row + 1):
                        cell = sheet.cell(row=r, column=col)
                        if cell.value is not None:
                            has_data = True
                            max_col = col
                            break
                    if not has_data:
                        break

                max_row = start_row
                for row in range(start_row+1, effective_max_row + 1):
                    has_data = False
                    for col in range(start_col, max_col + 1):
                        cell = sheet.cell(row=row, column=col)
                        if cell.value is not None:
                            has_data = True
                            max_row = row
                            break
                    if not has_data and row != start_row+1:
                        break

                # Step 1.5: Expand left to include additional columns
                for col in range(start_col - 1, 0, -1):  # Go left from start_col to column 1
                    has_data = False
                    for row in range(start_row, max_row + 1):  # Check within rectangular region
                        cell = sheet.cell(row=row, column=col)
                        if cell.value is not None:
                            has_data = True
                            start_col = col  # Update start_col to include this column
                            break
                    if not has_data:
                        break  # Found empty column, stop expanding left

                column_count = max_col - start_col + 1
                self.logger.info(f"Table boundaries: rows [{start_row}-{max_row}], cols [{start_col}-{max_col}], column_count={column_count}")

                # Step 2: Extract first few rows for header detection
                first_rows = []
                for row_idx in range(start_row, min(start_row + MAX_HEADER_DETECTION_ROWS, max_row + 1)):
                    row_values = []
                    for col in range(start_col, max_col + 1):
                        cell = sheet.cell(row=row_idx, column=col)
                        cell_data = self._process_cell(cell, None, row_idx, col)
                        row_values.append(cell_data["value"])
                    first_rows.append(row_values)

                # Step 3: Detect headers with LLM
                detection = await self.detect_excel_headers_with_llm(first_rows, llm)
                self.logger.info(f"Header detection result: has_headers={detection.has_headers}, num_header_rows={detection.num_header_rows}, confidence={detection.confidence}")

                # Step 4: Determine headers and data start row
                headers = []
                data_start_row = start_row

                if detection.has_headers and detection.num_header_rows == 1:
                    # Single row header - use directly
                    headers = first_rows[0] if first_rows else []
                    data_start_row = start_row + 1
                    self.logger.info(f"Using single-row headers: {headers}")
                elif detection.has_headers and detection.num_header_rows > 1:
                    # Multi-row headers: concatenate them into single-row headers
                    multirow_headers = first_rows[:detection.num_header_rows]
                    data_start_row = start_row + detection.num_header_rows
                    self.logger.info(f"Multi-row headers detected ({detection.num_header_rows} rows), will concatenate into single-row headers")

                    # Concatenate multi-row headers directly (no LLM)
                    headers = self._concatenate_multirow_headers(multirow_headers, column_count)
                    self.logger.info(f"Concatenated headers: {headers}")
                else:
                    # No headers: all rows are data, generate headers from data
                    data_start_row = start_row
                    sample_start = start_row
                    self.logger.info("No headers detected, will generate headers from data")

                    # Extract a bounded subset of rows for sampling to avoid scanning massive tables.
                    all_rows = []
                    sample_end_row = min(max_row, sample_start + EXCEL_HEADER_GENERATION_SAMPLE_SCAN_LIMIT - 1)
                    for row_idx in range(sample_start, sample_end_row + 1):
                        row_values = []
                        for col in range(start_col, max_col + 1):
                            cell = sheet.cell(row=row_idx, column=col)
                            cell_data = self._process_cell(cell, None, row_idx, col)
                            row_values.append(cell_data["value"])
                        all_rows.append(row_values)

                    # Select representative sample rows
                    sample_rows = self._select_representative_sample_rows(all_rows, MAX_HEADER_GENERATION_ROWS)
                    self.logger.info(f"Selected {len(sample_rows)} representative sample rows for header generation")

                    # Generate headers with LLM
                    headers = await self.generate_excel_headers_with_llm(sample_rows, column_count, llm)
                    self.logger.info(f"Generated headers: {headers}")

                # Normalize headers to match column count
                if headers:
                    # Replace None values in existing headers
                    for i in range(len(headers)):
                        if headers[i] is None or (isinstance(headers[i], str) and not headers[i].strip()):
                            headers[i] = f"Column_{i + 1}"

                    # Pad if too short
                    if len(headers) < column_count:
                        self.logger.info(f"Padding headers from {len(headers)} to {column_count}")
                        for i in range(len(headers) + 1, column_count + 1):
                            headers.append(f"Column_{i}")

                    # Truncate if too long (edge case)
                    elif len(headers) > column_count:
                        self.logger.warning(f"Truncating headers from {len(headers)} to {column_count}")
                        headers = headers[:column_count]

                    self.logger.info(f"Normalized headers ({len(headers)} total): {headers}")

                # Handle empty headers case
                if not headers or all(h is None for h in headers):
                    return {
                        "headers": [],
                        "data": [],
                        "start_row": start_row,
                        "start_col": start_col,
                        "end_row": start_row,
                        "end_col": start_col,
                    }

                # Step 5: Build table structure once with correct headers
                total_data_rows = max(0, max_row - data_start_row + 1)
                processed_end_row = min(max_row, data_start_row + EXCEL_MAX_TABLE_ROWS_TO_INDEX - 1)

                if total_data_rows > EXCEL_MAX_TABLE_ROWS_TO_INDEX:
                    self.logger.warning(
                        f"Large Excel table detected with {total_data_rows} rows at rows [{start_row}-{max_row}]. "
                        f"Limiting indexed rows to first {EXCEL_MAX_TABLE_ROWS_TO_INDEX}."
                    )

                table_data = []
                for row_idx in range(data_start_row, processed_end_row + 1):
                    row_data = []
                    for col_idx, col in enumerate(range(start_col, max_col + 1)):
                        cell = sheet.cell(row=row_idx, column=col)
                        header = headers[col_idx]
                        cell_data = self._process_cell(cell, header, row_idx, col)
                        row_data.append(cell_data)
                    table_data.append(row_data)

                return {
                    "headers": headers,
                    "data": table_data,
                    "start_row": start_row,
                    "start_col": start_col,
                    "end_row": processed_end_row,
                    "end_col": max_col,
                    "full_end_row": max_row,
                    "total_data_rows": total_data_rows,
                }

            # Find all tables in the sheet while skipping previously claimed ranges.
            row = 1
            while row <= effective_max_row:
                col = 1
                while col <= effective_max_col:
                    visited_range = get_visited_range(row, col)
                    if visited_range is not None:
                        _start_row, end_row, start_col, end_col = visited_range
                        if start_col == 1 and end_col >= effective_max_col and col == 1:
                            row = end_row + 1
                            col = 1
                            break
                        col = end_col + 1
                        continue

                    # O(1) set lookup instead of sheet.cell() creation
                    if (row, col) in non_empty_positions:
                        self.logger.info(f"Found potential table start at ({row}, {col})")
                        table = await get_table(row, col)
                        if table["data"]:  # Only add if table has data
                            tables.append(table)
                            full_end_row = table.get("full_end_row", table["end_row"])
                            visited_ranges.append(
                                (table["start_row"], full_end_row, table["start_col"], table["end_col"])
                            )
                            self.logger.info(f"Table added with {len(table['data'])} rows and {len(table['headers'])} columns")

                            if table["start_col"] == 1 and table["end_col"] >= effective_max_col:
                                row = full_end_row + 1
                                col = 1
                                break

                            col = table["end_col"] + 1
                            continue

                    col += 1
                else:
                    row += 1
                    continue

            self.logger.info(f"Found {len(tables)} tables in sheet: {sheet.title}")
            return tables

        except Exception as e:
            self.logger.error(f"Error finding tables in sheet {sheet.title}: {e}", exc_info=True)
            raise

    def _process_cell(self, cell: Cell | MergedCell, header: str | None, row: int, col: int) -> dict[str, Any]:
        """Process a single cell and return its data with denormalized merged cell values."""
        try:
            # Check if the cell is a merged cell
            if isinstance(cell, MergedCell):
                # Look for the merged range that contains this cell.
                merged_value = None

                for merged_range in cell.parent.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Get the top-left cell of the merged range
                        top_left_cell = cell.parent.cell(
                            row=merged_range.min_row, column=merged_range.min_col
                        )
                        merged_value = top_left_cell.value
                        # Apply datetime formatting if applicable
                        if isinstance(merged_value, datetime) and hasattr(top_left_cell, 'number_format'):
                            merged_value = format_excel_datetime(merged_value, top_left_cell.number_format)
                        break

                return {
                    "value": merged_value,  # Use the top-left cell's value (formatted if datetime)
                    "header": header,
                    "row": row,
                    "column": col,
                    "column_letter": get_column_letter(col),
                    "coordinate": f"{get_column_letter(col)}{row}",
                    "data_type": "merged",
                    "style": {"font": {}, "fill": {}, "alignment": {}},
                }

            # If not a merged cell, process normally.
            # Apply datetime formatting if the cell contains a datetime value
            cell_value = cell.value
            if isinstance(cell_value, datetime) and hasattr(cell, 'number_format'):
                cell_value = format_excel_datetime(cell_value, cell.number_format)

            return {
                "value": cell_value,  # Now contains formatted string for datetime values
                "header": header,
                "row": row,
                "column": col,
                "column_letter": cell.column_letter,
                "coordinate": cell.coordinate,
                "data_type": cell.data_type,
                "style": {
                    "font": {
                        "bold": cell.font.bold,
                        "italic": cell.font.italic,
                        "size": cell.font.size,
                        "color": cell.font.color.rgb if cell.font.color else None,
                    },
                    "fill": {
                        "background_color": (
                            cell.fill.start_color.rgb if cell.fill.start_color else None
                        )
                    },
                    "alignment": {
                        "horizontal": cell.alignment.horizontal,
                        "vertical": cell.alignment.vertical,
                    },
                },
            }
        except Exception as e:
            self.logger.error(f"Error processing cell at ({row}, {col}): {e}", exc_info=True)
            raise

    def _count_empty_values(self, row: dict[str, Any]) -> int:
        """Count the number of empty/None values in a row"""
        return sum(1 for value in row.values() if value is None or value == "")

    def _select_representative_sample_rows(
        self, data_rows: list[list[Any]], num_sample_rows: int = NUM_SAMPLE_ROWS
    ) -> list[tuple[int, list[Any], int]]:
        """
        Select representative sample rows from data by prioritizing rows with fewer empty values.

        This method selects up to num_sample_rows rows, prioritizing:
        1. Perfect rows with no empty values (stops early if enough are found)
        2. Rows with the fewest empty values as fallback

        Args:
            data_rows: List of rows (each row is a list of values)
            num_sample_rows: Number of sample rows to select (default: NUM_SAMPLE_ROWS)

        Returns:
            List of tuples (row_index, row_list, empty_count) sorted by original index
        """
        selected_rows = []
        fallback_rows = []

        for idx, row in enumerate(data_rows):
            # Count empty values in this row
            empty_count = sum(1 for value in row if value is None or (isinstance(value, str) and value.strip() == ""))

            if empty_count == 0:
                # Perfect row with no empty values
                selected_rows.append((idx, row, empty_count))
                if len(selected_rows) >= num_sample_rows:
                    break  # Early stop - found enough perfect rows
            else:
                # Keep track of best non-perfect rows as fallback
                fallback_rows.append((idx, row, empty_count))

        # If we didn't find enough perfect rows, supplement with the best fallback rows
        if len(selected_rows) < num_sample_rows:
            # Sort fallback rows by empty count (ascending), then by index
            fallback_rows.sort(key=lambda x: (x[2], x[0]))
            # Add the best fallback rows to reach the target count
            needed = num_sample_rows - len(selected_rows)
            selected_rows.extend(fallback_rows[:needed])

        # Sort by original index to maintain logical order
        selected_rows.sort(key=lambda x: x[0])

        return selected_rows

    def _convert_rows_to_strings(self, rows: list[list[Any]], num_rows: int = 4) -> list[list[str]]:
        """
        Convert multiple rows to lists of strings for LLM prompts.

        Args:
            rows: List of rows where each row is a list of values
            num_rows: Number of rows to convert (default: 4)

        Returns:
            List of converted rows, where each row is a list of strings.
            Returns empty list for rows that don't exist.
        """
        return [
            [str(v) if v is not None else "" for v in rows[i]] if i < len(rows) else []
            for i in range(num_rows)
        ]

    async def detect_excel_headers_with_llm(
        self, first_rows: list[list[Any]], llm: BaseChatModel
    ) -> ExcelHeaderDetection:
        """
        Use LLM to detect if the first row(s) contain valid headers and how many rows they span.

        Args:
            first_rows: List of first few rows as lists of values (typically 4-6 rows)
            llm: Language model instance

        Returns:
            ExcelHeaderDetection with has_headers, num_header_rows, confidence, and reasoning
        """
        self.logger.info(f"Detecting headers with LLM for {len(first_rows)} rows")
        try:
            # Handle edge case: no rows available
            if not first_rows:
                self.logger.warning("No rows available for header detection")
                return ExcelHeaderDetection(
                    has_headers=False,
                    num_header_rows=0,
                    confidence="high",
                    reasoning="No rows available in the table"
                )

            # Build dynamic row text based on actual available rows
            rows_text_lines = []
            for i, row in enumerate(first_rows[:4], start=1):  # Show up to 4 rows
                row_str = [str(v) if v is not None else "" for v in row]
                rows_text_lines.append(f"Row {i}: {row_str}")

            rows_text = "\n".join(rows_text_lines)

            self.logger.info("Calling LLM for header detection")
            messages = excel_header_detection_prompt.format_messages(
                rows_text=rows_text
            )

            # Use centralized utility with reflection
            parsed_response = await invoke_with_structured_output_and_reflection(
                llm, messages, ExcelHeaderDetection
            )

            if parsed_response is not None:
                self.logger.info(f"LLM header detection successful: {parsed_response.reasoning}")
                # Validate num_header_rows is sensible
                if parsed_response.num_header_rows < 0:
                    parsed_response.num_header_rows = 0

                # If has_headers is True but num_header_rows is 0, correct it
                if parsed_response.has_headers and parsed_response.num_header_rows == 0:
                    parsed_response.num_header_rows = 1

                # If has_headers is False, ensure num_header_rows is 0
                if not parsed_response.has_headers:
                    parsed_response.num_header_rows = 0

                return parsed_response

            self.logger.warning("Header detection LLM call failed, defaulting to no headers")
            return ExcelHeaderDetection(
                has_headers=False,
                num_header_rows=0,
                confidence="low",
                reasoning="LLM call failed, defaulting to no headers"
            )

        except Exception as e:
            self.logger.warning(f"Error in Excel header detection: {e}, defaulting to no headers")
            return ExcelHeaderDetection(
                has_headers=False,
                num_header_rows=0,
                confidence="low",
                reasoning=f"Error occurred: {str(e)}, defaulting to no headers"
            )

    async def generate_excel_headers_with_llm(
        self, sample_rows: list[tuple[int, list[Any], int]], column_count: int, llm: BaseChatModel
    ) -> list[str]:
        """
        Generate descriptive headers from sample data using LLM.

        Used for two scenarios:
        1. No headers detected - generate from data rows
        2. Multi-row headers detected - generate from all rows including the multi-row headers

        Args:
            sample_rows: List of tuples (row_index, row_list, empty_count) from _select_representative_sample_rows
            column_count: Number of columns expected
            llm: Language model instance

        Returns:
            List of generated header names (always exactly column_count items)
        """
        self.logger.info(f"Generating headers with LLM for {column_count} columns using {len(sample_rows)} sample rows")

        try:
            # Format sample data for display
            formatted_samples = []
            for _idx, row, _empty_count in sample_rows[:MAX_HEADER_GENERATION_ROWS]:
                formatted_row = [str(v) if v is not None else "" for v in row]
                formatted_samples.append(formatted_row)

            # Format as JSON string for prompt
            sample_data_str = json.dumps(formatted_samples, indent=2, ensure_ascii=False)

            # Initial messages
            messages = excel_header_generation_prompt.format_messages(
                sample_data=sample_data_str,
                column_count=column_count,
                sample_count=len(formatted_samples),
            )

            # Retry loop for count mismatches
            for attempt in range(MAX_HEADER_COUNT_RETRIES + 1):
                if attempt > 0:
                    self.logger.info(f"Retry attempt {attempt}/{MAX_HEADER_COUNT_RETRIES} for header generation")
                else:
                    self.logger.info("Calling LLM for header generation (initial attempt)")

                # Use centralized utility with reflection for parse errors
                parsed_response = await invoke_with_structured_output_and_reflection(
                    llm, messages, TableHeaders
                )

                if parsed_response is not None and parsed_response.headers:
                    generated_headers = parsed_response.headers
                    self.logger.info(f"LLM generated {len(generated_headers)} headers (expected {column_count})")

                    # Validate header count matches column count
                    if len(generated_headers) == column_count:
                        self.logger.info(f"Successfully generated headers matching expected column count ({column_count})")
                        return generated_headers
                    else:
                        self.logger.warning(
                            f"Header count mismatch: generated {len(generated_headers)}, expected {column_count}. "
                            f"Headers: {generated_headers}"
                        )

                        # If we have retries left, add reflection message for count correction
                        if attempt < MAX_HEADER_COUNT_RETRIES:
                            self.logger.info("Adding reflection message to correct header count")

                            # Convert messages to list if not already
                            messages_list = list(messages)

                            # Add the failed response to context
                            failed_response = json.dumps({"headers": generated_headers}, indent=2, ensure_ascii=False)
                            messages_list.append(AIMessage(content=failed_response))

                            # Add reflection prompt
                            reflection_prompt = f"""Your previous response contained {len(generated_headers)} headers, but I need EXACTLY {column_count} headers.

Previous headers you provided: {generated_headers}

ERROR: You returned {len(generated_headers)} headers but the data has {column_count} columns.

Please correct your response:
- Analyze the sample data again carefully
- Count that there are {column_count} columns in the data
- Generate EXACTLY {column_count} headers, one for each column
- Verify your count before responding

Respond with ONLY a JSON object with EXACTLY {column_count} headers:
{{
    "headers": ["Header1", "Header2", ..., "Header{column_count}"]
}}"""

                            messages_list.append(HumanMessage(content=reflection_prompt))
                            messages = messages_list
                            continue  # Try again
                        else:
                            # Out of retries, try smart fallback
                            self.logger.warning(f"Exhausted {MAX_HEADER_COUNT_RETRIES} retries")

                else:
                    self.logger.warning("LLM returned no response or empty headers")
                    break  # Exit retry loop

            # Final fallback: generate generic headers
            self.logger.warning(f"Using generic fallback headers for {column_count} columns")
            return [f"Column_{i}" for i in range(1, column_count + 1)]

        except Exception as e:
            self.logger.error(f"Error in Excel header generation: {e}", exc_info=True)
            self.logger.warning(f"Using generic fallback headers for {column_count} columns")
            return [f"Column_{i}" for i in range(1, column_count + 1)]




    def _concatenate_multirow_headers(self, multirow_headers: list[list[Any]], column_count: int) -> list[str]:
        """
        Fallback method to concatenate multi-row headers with underscores.

        Args:
            multirow_headers: List of header rows
            column_count: Expected number of columns

        Returns:
            List of concatenated header strings
        """
        self.logger.info(f"Using simple concatenation for {len(multirow_headers)} header rows")
        consolidated = []

        for col_idx in range(column_count):
            # Collect non-empty values from all header rows for this column
            parts = []
            seen = set()  # Track seen values to avoid duplicates (e.g., from merged cells)

            for header_row in multirow_headers:
                if col_idx < len(header_row):
                    value = header_row[col_idx]
                    if value is not None and str(value).strip():
                        value_str = str(value).strip()
                        # Only add if we haven't seen this value already (handles merged cells)
                        if value_str not in seen:
                            parts.append(value_str)
                            seen.add(value_str)

            # Join with underscores or use generic name if no parts
            header = "_".join(parts) if parts else f"Column_{col_idx + 1}"

            consolidated.append(header)

        return consolidated

    @retry(
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=1, min=1, max=10),
        before_sleep=lambda retry_state: retry_state.args[0].logger.warning(
            f"Retrying LLM call after error. Attempt {retry_state.attempt_number}"
        ),
    )
    async def _call_llm(self, messages: list[Any]) -> AIMessage:
        """Wrapper for LLM calls with retry logic"""
        return await self.llm.ainvoke(messages)  # type: ignore[return-value]

    async def get_tables_in_sheet(self, sheet_name: str, llm: BaseChatModel) -> list[dict[str, Any]]:
        """Get all tables in a specific sheet with LLM-based header detection/generation

        Note: Header detection and generation is now handled in find_tables() method,
        so this method simply returns the tables with properly detected/generated headers.
        """
        self.logger.info(f"Getting tables in sheet: {sheet_name}")
        try:
            if not self.workbook:
                raise ValueError("Workbook not loaded")

            if sheet_name not in self.workbook.sheetnames:
                self.logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                return []

            sheet = self.workbook[sheet_name]
            # find_tables now handles header detection/generation internally
            tables = await self.find_tables(sheet, llm)

            self.logger.info(f"Retrieved {len(tables)} tables from sheet: {sheet_name}")
            return tables

        except Exception as e:
            self.logger.error(f"Error getting tables in sheet {sheet_name}: {e}", exc_info=True)
            raise

    async def get_table_summary(self, table: dict[str, Any]) -> str:
        """Get a natural language summary of a specific table"""
        self.logger.info(f"Getting summary for table with {len(table['headers'])} columns and {len(table['data'])} rows")
        try:
            # Prepare sample data
            sample_data = [
                {
                    cell["header"]: (
                        cell["value"].isoformat()
                        if isinstance(cell["value"], datetime)
                        else cell["value"]
                    )
                    for cell in row
                }
                for row in table["data"][:3]  # Use first 3 rows as sample
            ]

            # Get summary from LLM with retry
            messages = self.table_summary_prompt.format_messages(
                headers=table["headers"], sample_data=json.dumps(sample_data, indent=2, ensure_ascii=False)
            )
            response = await self._call_llm(messages)
            # Gemini and similar return content as a list of blocks, not a string.
            summary = coerce_message_content_to_text(response.content)
            if '</think>' in summary:
                summary = summary.split('</think>')[-1]
            self.logger.info("Table summary generated")
            return summary

        except Exception as e:
            self.logger.error(f"Error getting table summary: {e}", exc_info=True)
            raise

    async def get_rows_text(
        self, rows: list[list[dict[str, Any]]], table_summary: str
    ) -> list[str]:
        """Convert multiple rows into natural language text using context from summaries in a single prompt"""
        self.logger.info(f"Converting {len(rows)} rows to natural language text")
        try:
            # Prepare rows data
            rows_data = [
                {
                    cell["header"]: (
                        cell["value"].isoformat()
                        if isinstance(cell["value"], datetime)
                        else cell["value"]
                    )
                    for cell in row
                }
                for row in rows
            ]

            # Get natural language text from LLM with count validation
            # Use CSV prompt which includes explicit row count validation
            messages = row_text_prompt_for_csv.format_messages(
                table_summary=table_summary,
                numbered_rows_data=format_rows_with_index(rows_data),
                row_count=len(rows_data)
            )

            # Default to simple text representations of rows
            descriptions = [generate_simple_row_text(row) for row in rows_data]

            # Use centralized utility with reflection and count validation
            parsed_response = await invoke_with_row_descriptions_and_reflection(
                self.llm, messages, expected_count=len(rows_data)
            )

            if parsed_response is not None and parsed_response.descriptions:
                descriptions = parsed_response.descriptions
                self.logger.info(f"Successfully generated natural language descriptions for {len(descriptions)} rows")
            else:
                self.logger.warning(f"LLM failed to generate descriptions, using fallback for {len(rows_data)} rows")

            return descriptions
        except Exception as e:
            self.logger.error(f"Error converting rows to natural language text: {e}", exc_info=True)
            raise

    async def process_sheet_with_summaries(
        self, llm: BaseChatModel, sheet_name: str, cumulative_row_count: list[int]
    ) -> dict[str, Any] | None:
        """Process a sheet and generate all summaries and row texts
        Args:
            llm: Language model instance
            sheet_name: Name of the sheet to process
            cumulative_row_count: List with single element [count] to track cumulative rows across all tables
        """
        self.logger.info(f"Processing sheet with summaries: {sheet_name}")
        self.llm = llm

        if not self.workbook:
            self.logger.warning(f"Workbook not loaded, cannot process sheet '{sheet_name}'")
            return None

        if sheet_name not in self.workbook.sheetnames:
            self.logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            return None

        # Get threshold from environment variable (default: 1000)
        threshold = int(os.getenv("MAX_TABLE_ROWS_FOR_LLM", "1000"))
        self.logger.info(f"Using LLM threshold for row processing: {threshold} (cumulative count: {cumulative_row_count[0]})")

        # Get tables in the sheet
        tables = await self.get_tables_in_sheet(sheet_name, llm)

        # Process each table
        processed_tables = []
        for table_idx, table in enumerate(tables, 1):
            self.logger.info(f"Processing table {table_idx}/{len(tables)} in sheet {sheet_name}")
            # Get table summary (always use LLM)
            table_summary = await self.get_table_summary(table)

            # Add current table rows to cumulative count
            table_row_count = len(table["data"])
            cumulative_row_count[0] += table_row_count
            self.logger.info(f"Table has {table_row_count} rows, cumulative count: {cumulative_row_count[0]}")

            # Check if cumulative count exceeds threshold
            use_llm_for_rows = cumulative_row_count[0] <= threshold

            processed_rows = []

            if use_llm_for_rows:
                self.logger.info(f"Using LLM for row processing (under threshold of {threshold})")
                # Process rows in batches of 50 in parallel using LLM
                batch_size = 50

                # Create batches
                batches = []
                for i in range(0, len(table["data"]), batch_size):
                    batch = table["data"][i : i + batch_size]
                    batches.append((i, batch))  # Store start index and batch data

                self.logger.info(f"Processing {len(table['data'])} rows in {len(batches)} batches of {batch_size}")

                # Limit parallel processing to at most 10 concurrent batches
                semaphore = asyncio.Semaphore(10)

                async def limited_get_rows_text(batch: list[Any], _sem: asyncio.Semaphore = semaphore, _ts: str = table_summary) -> list[str]:
                    async with _sem:
                        return await self.get_rows_text(batch, _ts)

                # Create throttled tasks for all batches
                batch_tasks = []
                for start_idx, batch in batches:
                    task = limited_get_rows_text(batch)
                    batch_tasks.append((start_idx, batch, task))

                # Wait for all batches to complete (max 10 running concurrently)
                task_results = await asyncio.gather(*[task for _, _, task in batch_tasks])
                self.logger.info(f"Completed processing {len(batch_tasks)} batches with LLM")

                # Combine results with their metadata and process
                for i, (_start_idx, batch, _) in enumerate(batch_tasks):
                    row_texts = task_results[i]

                    # Add processed rows to results
                    for row, row_text in zip(batch, row_texts):
                        if row:
                            processed_rows.append(
                                {
                                    "raw_data": {cell["header"]: cell["value"] for cell in row},
                                    "natural_language_text": row_text,
                                    "row_num": row[0]["row"],  # Include row number
                                }
                            )
            else:
                self.logger.info(f"Using simple format for row processing (exceeded threshold of {threshold})")
                # Use simple format for rows (skip LLM)
                for row in table["data"]:
                    if row:
                        row_data = {cell["header"]: cell["value"] for cell in row}
                        row_text = generate_simple_row_text(row_data)
                        processed_rows.append(
                            {
                                "raw_data": row_data,
                                "natural_language_text": row_text,
                                "row_num": row[0]["row"]  # Include row number
                            }
                        )

            processed_tables.append(
                {
                    "headers": table["headers"],
                    "summary": table_summary,
                    "rows": processed_rows,
                    "location": {
                        "start_row": table["start_row"],
                        "start_col": table["start_col"],
                        "end_row": table["end_row"],
                        "end_col": table["end_col"],
                    },
                }
            )
            self.logger.info(f"Completed processing table {table_idx} with {len(processed_rows)} rows")

        self.logger.info(f"Completed processing sheet {sheet_name} with {len(processed_tables)} tables")
        return {"sheet_name": sheet_name, "tables": processed_tables}

    async def get_blocks_from_workbook(self, llm: BaseChatModel) -> BlocksContainer:
        """Build a BlocksContainer with SHEET and TABLE groups and TABLE_ROW blocks.

        Mirrors the CSV blocks structure, but nests tables under sheet groups.
        """
        self.logger.info("Building blocks from workbook")
        blocks: list[Block] = []
        block_groups: list[BlockGroup] = []

        # Initialize cumulative row count for record-level threshold checking
        cumulative_row_count = [0]

        # Iterate sheets and build hierarchy
        assert self.workbook is not None, "Workbook must be loaded before calling get_blocks_from_workbook"
        self.logger.info(f"Processing {len(self.workbook.sheetnames)} sheets: {self.workbook.sheetnames}")
        for sheet_idx, sheet_name in enumerate(self.workbook.sheetnames, 1):
            self.logger.info(f"Processing sheet {sheet_idx}/{len(self.workbook.sheetnames)}: {sheet_name}")
            sheet_result = await self.process_sheet_with_summaries(llm, sheet_name, cumulative_row_count)
            if sheet_result is None:
                continue

            # Create SHEET group
            sheet_group_index = len(block_groups)
            sheet_table_group_indices = []
            sheet_group = BlockGroup(
                index=sheet_group_index,
                name=sheet_result["sheet_name"],
                type=GroupType.SHEET,
                parent_index=None,
                description=None,
                table_metadata=None,
                data={
                    "sheet_name": sheet_result["sheet_name"],
                    "table_count": len(sheet_result["tables"]),
                },
                format=DataFormat.JSON,
            )
            block_groups.append(sheet_group)

            # Add TABLE groups under this sheet
            for table in sheet_result["tables"]:
                table_group_index = len(block_groups)

                headers = table.get("headers", [])
                rows = table.get("rows", [])

                table_row_block_indices = []
                num_of_rows = len(rows)
                num_of_cols = len(headers) if headers else (len(rows[0]["raw_data"]) if rows else 0)
                num_of_cells = num_of_rows * num_of_cols
                table_group = BlockGroup(
                    index=table_group_index,
                    name=None,
                    type=GroupType.TABLE,
                    parent_index=sheet_group_index,
                    description=None,
                    source_group_id=None,
                    table_metadata=TableMetadata(
                        num_of_rows=num_of_rows,
                        num_of_cols=num_of_cols,
                        num_of_cells=num_of_cells,
                    ),
                    data={
                        "table_summary": table.get("summary", ""),
                        "column_headers": headers,
                        "sheet_number": sheet_idx,
                        "sheet_name": sheet_name,
                    },
                    format=DataFormat.JSON,
                )
                block_groups.append(table_group)
                sheet_table_group_indices.append(table_group_index)


                # Create TABLE_ROW blocks under this table
                for i, row in enumerate(rows):
                    block_index = len(blocks)
                    blocks.append(
                        Block(
                            index=block_index,
                            type=BlockType.TABLE_ROW,
                            format=DataFormat.JSON,
                            data={
                                "row_natural_language_text": row.get("natural_language_text", ""),
                                "row_number": int(row.get("row_num") or (i + 1)),
                                "row_end_number": int(row.get("row_end_num") or row.get("row_num") or (i + 1)),
                                "row_count": int(row.get("row_count") or 1),
                                "sheet_number": sheet_idx,
                                "sheet_name": sheet_name,
                            },
                            parent_index=table_group_index,
                        )
                    )
                    table_row_block_indices.append(block_index)

                # attach table children using range-based structure
                block_groups[table_group_index].children = BlockGroupChildren.from_indices(
                    block_indices=table_row_block_indices
                )

            # attach sheet children (its tables) using range-based structure
            block_groups[sheet_group_index].children = BlockGroupChildren.from_indices(
                block_group_indices=sheet_table_group_indices
            )

        self.logger.info(f"Workbook processing complete. Total: {len(blocks)} blocks, {len(block_groups)} block groups")
        return BlocksContainer(blocks=blocks, block_groups=block_groups)

