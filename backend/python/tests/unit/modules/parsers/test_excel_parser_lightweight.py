"""Unit tests for ExcelParser.create_blocks_lightweight."""

import io
from unittest.mock import MagicMock

import pytest
from openpyxl import Workbook

from app.models.blocks import BlockType, GroupType
from app.modules.parsers.excel.excel_parser import ExcelParser


def _xlsx_bytes(rows: list[list[object]], sheet_name: str = "Sheet1") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def parser():
    return ExcelParser(logger=MagicMock(), config_service=MagicMock())


@pytest.mark.asyncio
async def test_create_blocks_lightweight_happy_path(parser):
    content = _xlsx_bytes([["name", "amount"], ["Alice", 10], ["Bob", 20]])
    parser.load_workbook_from_binary(content)
    container = await parser.create_blocks_lightweight()

    assert any(g.type == GroupType.SHEET for g in container.block_groups)
    assert any(g.type == GroupType.TABLE for g in container.block_groups)
    assert len([b for b in container.blocks if b.type == BlockType.TABLE_ROW]) == 2


@pytest.mark.asyncio
async def test_create_blocks_lightweight_respects_max_rows(parser):
    rows = [["col"]] + [[i] for i in range(50)]
    content = _xlsx_bytes(rows)
    parser.load_workbook_from_binary(content)
    container = await parser.create_blocks_lightweight(max_rows=5)
    assert len([b for b in container.blocks if b.type == BlockType.TABLE_ROW]) == 5


@pytest.mark.asyncio
async def test_create_blocks_lightweight_empty_workbook(parser):
    content = _xlsx_bytes([])  # header-less empty sheet
    parser.load_workbook_from_binary(content)
    container = await parser.create_blocks_lightweight()
    # Empty / header-only sheet yields no data rows.
    assert container.blocks == []


@pytest.mark.asyncio
async def test_create_blocks_lightweight_closes_workbook(parser):
    content = _xlsx_bytes([["h"], ["v"]])
    parser.load_workbook_from_binary(content)
    wb = parser.workbook
    close = MagicMock(wraps=wb.close)
    wb.close = close
    await parser.create_blocks_lightweight()
    close.assert_called_once()
